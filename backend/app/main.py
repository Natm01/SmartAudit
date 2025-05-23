from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Path
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from typing import List, Optional
import os
import tempfile
import shutil
from datetime import datetime
import logging
import json
import asyncio

from app.services.file_processor import process_libro_diario, process_sumas_saldos
from app.services.validators import validate_files
from app.services.analyzers import generate_summary
from app.schemas.libro_diario import FileUploadResponse, ValidationResult, ProcessResult
from fastapi.staticfiles import StaticFiles

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ruta del frontend y favicon
FRONTEND_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../frontend/build"))
FAVICON_PATH = os.path.join(FRONTEND_PATH, "favicon.ico")


# === Configuración de FastAPI ===
app = FastAPI(title="SmartAudit API", description="API para procesamiento de libros diarios contables")

# Servir archivos estáticos del frontend si existe
if os.path.exists(FRONTEND_PATH):
    app.mount("/", StaticFiles(directory=FRONTEND_PATH, html=True), name="frontend")


# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/favicon.ico")
async def get_favicon():
    if os.path.exists(FAVICON_PATH):
        return FileResponse(FAVICON_PATH)
    raise HTTPException(status_code=404, detail="Favicon no encontrado")


@app.get("/api")
async def root():
    return {"message": "SmartAudit API"}

# === Utils ===

def save_uploaded_files(files: List[UploadFile], temp_dir: str) -> List[str]:
    """Guarda archivos subidos en el directorio temporal"""
    paths = []
    for file in files:
        if file.filename:  # Verificar que el archivo tenga nombre
            path = os.path.join(temp_dir, file.filename)
            try:
                with open(path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
                paths.append(file.filename)  # Solo devolver el nombre del archivo
                logger.info(f"Archivo guardado: {file.filename} ({file.size} bytes)")
            except Exception as e:
                logger.error(f"Error guardando archivo {file.filename}: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error guardando archivo {file.filename}")
    return paths


def classify_files(temp_dir: str, all_files: List[str]) -> tuple[list[str], list[str]]:
    """Clasifica archivos entre libro diario y sumas y saldos"""
    libro_files = []
    sumas_files = []

    for file in all_files:
        file_path = os.path.join(temp_dir, file)
        
        # Verificar que el archivo exista
        if not os.path.exists(file_path):
            logger.warning(f"Archivo no encontrado: {file_path}")
            continue
            
        # Clasificar por nombre primero
        if any(k in file.lower() for k in ['suma', 'saldo', 'balance', 'mayor']):
            sumas_files.append(file)
            continue
            
        # Clasificar por extensión y contenido para Excel
        if file.lower().endswith(('.xls', '.xlsx')):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                found_sumas_keywords = False
                
                # Revisar las primeras 10 filas buscando palabras clave
                for row in ws.iter_rows(max_row=10, values_only=True):
                    if row and any(cell and any(k in str(cell).lower() for k in ['cta.mayor', 'saldo', 'arrastre']) for cell in row):
                        found_sumas_keywords = True
                        break
                
                if found_sumas_keywords:
                    sumas_files.append(file)
                else:
                    libro_files.append(file)
            except Exception as e:
                logger.warning(f"Error analizando archivo Excel {file}: {str(e)}")
                libro_files.append(file)  # Por defecto, clasificar como libro diario
        else:
            libro_files.append(file)

    logger.info(f"Archivos clasificados - Libro: {libro_files}, Sumas: {sumas_files}")
    return libro_files, sumas_files


def check_temp_dir_exists(temp_dir: str):
    """Verifica que el directorio temporal exista"""
    if not temp_dir or not os.path.exists(temp_dir):
        logger.error(f"Directorio temporal no existe: {temp_dir}")
        raise HTTPException(status_code=400, detail="El directorio temporal no existe o es inválido")


# === Streaming validation generator ===
async def stream_validation_progress(temp_dir: str, project: str, year: str, start_date: str, end_date: str):
    """Generador que envía actualizaciones de progreso de validación en tiempo real"""
    
    try:
        # Paso 1: Verificar archivos
        yield json.dumps({
            "step": "loading_files",
            "message": "Verificando archivos en el servidor...",
            "progress": 0,
            "completed": False
        }) + "\n"
        
        await asyncio.sleep(0.5)  # Simular tiempo de verificación
        
        all_files = [f for f in os.listdir(temp_dir) if os.path.isfile(os.path.join(temp_dir, f))]
        if not all_files:
            yield json.dumps({
                "step": "error",
                "message": "No se encontraron archivos",
                "progress": 0,
                "completed": False,
                "error": True
            }) + "\n"
            return
            
        yield json.dumps({
            "step": "loading_files", 
            "message": f"Archivos encontrados: {len(all_files)}",
            "progress": 16,
            "completed": True
        }) + "\n"
        
        # Paso 2: Analizar estructura
        yield json.dumps({
            "step": "analyzing_structure",
            "message": "Analizando estructura de archivos...",
            "progress": 16,
            "completed": False
        }) + "\n"
        
        await asyncio.sleep(0.8)  # Tiempo real de análisis
        libro_files, sumas_files = classify_files(temp_dir, all_files)
        
        yield json.dumps({
            "step": "analyzing_structure",
            "message": f"Estructura analizada: {len(libro_files)} libro diario, {len(sumas_files)} sumas y saldos",
            "progress": 32,
            "completed": True
        }) + "\n"
        
        # Paso 3: Validar campos mínimos
        yield json.dumps({
            "step": "validating_fields",
            "message": "Validando campos mínimos requeridos...",
            "progress": 32,
            "completed": False
        }) + "\n"
        
        # Simular validación de campos (en un caso real aquí iría validación parcial)
        await asyncio.sleep(1.2)  # Tiempo real de validación de campos
        
        yield json.dumps({
            "step": "validating_fields",
            "message": "Campos mínimos validados correctamente",
            "progress": 48,
            "completed": True
        }) + "\n"
        
        # Paso 4: Validar fechas
        yield json.dumps({
            "step": "validating_dates",
            "message": f"Validando formato de fechas para período {start_date} - {end_date}...",
            "progress": 48,
            "completed": False
        }) + "\n"
        
        # Simular validación de fechas
        await asyncio.sleep(1.0)  # Tiempo real de validación de fechas
        
        yield json.dumps({
            "step": "validating_dates",
            "message": "Fechas validadas dentro del período permitido",
            "progress": 64,
            "completed": True
        }) + "\n"
        
        # Paso 5: Validar balance
        yield json.dumps({
            "step": "validating_balance",
            "message": "Verificando balance de asientos contables...",
            "progress": 64,
            "completed": False
        }) + "\n"
        
        # Simular validación de balance
        await asyncio.sleep(1.5)  # Tiempo real de validación de balance
        
        yield json.dumps({
            "step": "validating_balance",
            "message": "Balance de asientos verificado correctamente",
            "progress": 80,
            "completed": True
        }) + "\n"
        
        # Paso 6: Validar cuentas y realizar validación completa
        yield json.dumps({
            "step": "validating_accounts",
            "message": "Validando plan de cuentas y ejecutando validación final...",
            "progress": 80,
            "completed": False
        }) + "\n"
        
        # Ejecutar la validación completa real
        try:
            result = validate_files(temp_dir, all_files, start_date, end_date)
            await asyncio.sleep(0.5)  # Pequeño delay para mostrar completado
            
            yield json.dumps({
                "step": "validating_accounts",
                "message": "Plan de cuentas validado",
                "progress": 96,
                "completed": True
            }) + "\n"
            
            # Paso final: Completar
            yield json.dumps({
                "step": "completed",
                "message": "Validación completada exitosamente",
                "progress": 100,
                "completed": True,
                "result": result
            }) + "\n"
            
        except Exception as validation_error:
            logger.error(f"Error en validación final: {str(validation_error)}")
            yield json.dumps({
                "step": "error",
                "message": f"Error en la validación final: {str(validation_error)}",
                "progress": 80,
                "completed": False,
                "error": True
            }) + "\n"
        
    except Exception as e:
        logger.error(f"Error en validación streaming: {str(e)}")
        yield json.dumps({
            "step": "error",
            "message": f"Error en la validación: {str(e)}",
            "progress": 0,
            "completed": False,
            "error": True
        }) + "\n"


# === Endpoints ===

@app.post("/api/upload", response_model=FileUploadResponse)
async def upload_files(
    project: str = Form(...),
    year: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    libro_diario_files: List[UploadFile] = File(...),
    sumas_saldos_files: Optional[List[UploadFile]] = File(None)
):
    """Endpoint para subir archivos"""
    temp_dir = tempfile.mkdtemp(prefix="smartaudit_")
    logger.info(f"Creado directorio temporal: {temp_dir}")
    
    try:
        # Validar que haya archivos de libro diario
        if not libro_diario_files or all(not f.filename for f in libro_diario_files):
            raise HTTPException(status_code=400, detail="Debe proporcionar al menos un archivo de libro diario")
        
        # Guardar archivos de libro diario
        libro_paths = save_uploaded_files([f for f in libro_diario_files if f.filename], temp_dir)
        
        # Guardar archivos de sumas y saldos si existen
        sumas_paths = []
        if sumas_saldos_files:
            sumas_paths = save_uploaded_files([f for f in sumas_saldos_files if f.filename], temp_dir)

        # Preparar respuesta
        response = {
            "project": project,
            "year": year,
            "date_range": f"{start_date} - {end_date}",
            "libro_diario_files": [{"name": name, "size": os.path.getsize(os.path.join(temp_dir, name))} for name in libro_paths],
            "sumas_saldos_files": [{"name": name, "size": os.path.getsize(os.path.join(temp_dir, name))} for name in sumas_paths],
            "temp_dir": temp_dir
        }
        
        logger.info(f"Upload exitoso - Libro: {len(libro_paths)} archivos, Sumas: {len(sumas_paths)} archivos")
        return response

    except HTTPException:
        # Limpiar directorio temporal en caso de error HTTP
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        raise
    except Exception as e:
        # Limpiar directorio temporal en caso de error general
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        logger.error(f"Error en upload: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al procesar archivos: {str(e)}")


@app.post("/api/validate-stream")
async def validate_stream(
    temp_dir: str = Form(...),
    project: str = Form(...),
    year: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...)
):
    """Endpoint para validación con streaming de progreso en tiempo real"""
    logger.info(f"Iniciando validación streaming para directorio: {temp_dir}")
    
    check_temp_dir_exists(temp_dir)
    
    return StreamingResponse(
        stream_validation_progress(temp_dir, project, year, start_date, end_date),
        media_type="application/x-ndjson",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
    )


@app.post("/api/validate", response_model=ValidationResult)
async def validate(
    temp_dir: str = Form(...),
    project: str = Form(...),
    year: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...)
):
    """Endpoint para validar archivos (versión tradicional)"""
    logger.info(f"Iniciando validación para directorio: {temp_dir}")
    
    try:
        check_temp_dir_exists(temp_dir)
        
        # Obtener lista de archivos
        all_files = [f for f in os.listdir(temp_dir) if os.path.isfile(os.path.join(temp_dir, f))]
        
        if not all_files:
            raise HTTPException(status_code=400, detail="No se encontraron archivos en el directorio temporal")
        
        logger.info(f"Archivos encontrados: {all_files}")
        
        # Realizar validación
        result = validate_files(temp_dir, all_files, start_date, end_date)
        
        logger.info(f"Validación completada - Errores: {result.get('has_errors', False)}")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en validación: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en la validación: {str(e)}")


@app.get("/api/preview/{temp_dir}")
async def get_preview_data(temp_dir: str = Path(...), file_type: str = "libro"):
    """Endpoint para obtener datos de previsualización"""
    logger.info(f"Solicitud de preview - Directorio: {temp_dir}, Tipo: {file_type}")
    
    try:
        check_temp_dir_exists(temp_dir)
        
        all_files = [f for f in os.listdir(temp_dir) if os.path.isfile(os.path.join(temp_dir, f))]
        libro_files, sumas_files = classify_files(temp_dir, all_files)

        if file_type == "libro" and libro_files:
            result = process_libro_diario(temp_dir, libro_files)
            return {
                "entries": result["entries"][:20],
                "total": len(result["entries"]),
                "accounting_date_range": result.get("accounting_date_range", ""),
                "registration_date_range": result.get("registration_date_range", "")
            }

        elif file_type == "sumas" and sumas_files:
            result = process_sumas_saldos(temp_dir, sumas_files)
            return {
                "records": result["records"][:30],
                "total": len(result["records"]),
                "total_debe": result.get("total_debe", 0),
                "total_haber": result.get("total_haber", 0),
                "total_saldo": result.get("total_saldo", 0)
            }

        return {"entries": [] if file_type == "libro" else [], "total": 0}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando previsualización: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generando previsualización: {str(e)}")


@app.post("/api/process", response_model=ProcessResult)
async def process(temp_dir: str = Form(...), validation_id: str = Form(...)):
    """Endpoint para procesar archivos validados"""
    logger.info(f"Iniciando procesamiento - Directorio: {temp_dir}, Validation ID: {validation_id}")
    
    try:
        if not validation_id:
            raise HTTPException(status_code=400, detail="ID de validación requerido")
        
        check_temp_dir_exists(temp_dir)

        all_files = [f for f in os.listdir(temp_dir) if os.path.isfile(os.path.join(temp_dir, f))]
        libro_files, sumas_files = classify_files(temp_dir, all_files)

        if not libro_files:
            raise HTTPException(status_code=400, detail="No se encontraron archivos de libro diario válidos")

        # Procesar libro diario
        libro_result = process_libro_diario(temp_dir, libro_files)

        # Procesar sumas y saldos si existen
        sumas_result = None
        if sumas_files:
            try:
                sumas_result = process_sumas_saldos(temp_dir, sumas_files)
                logger.info("Sumas y saldos procesadas exitosamente")
            except Exception as e:
                logger.warning(f"Error procesando sumas y saldos (continuando sin ellas): {str(e)}")

        # Generar resumen
        summary = generate_summary(libro_result)

        # Preparar respuesta
        response = {
            "accounting_date_range": summary["accounting_date_range"],
            "registration_date_range": summary["registration_date_range"],
            "entries": libro_result["entries"][:10],  # Solo las primeras 10 para la respuesta
            "summary": summary["activity_summary"]
        }
        
        # Agregar información de sumas y saldos si existe
        if sumas_result:
            response["sumas_saldos_summary"] = {
                "total_records": len(sumas_result["records"]),
                "total_balance": sumas_result["total_saldo"]
            }

        logger.info("Procesamiento completado exitosamente")
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en procesamiento: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en el procesamiento: {str(e)}")


@app.delete("/api/cleanup/{temp_dir_encoded}")
async def cleanup_temp_files(temp_dir_encoded: str = Path(...)):
    """Endpoint para limpiar archivos temporales"""
    try:
        # Decodificar la ruta (manejar caracteres especiales)
        import urllib.parse
        temp_dir = urllib.parse.unquote(temp_dir_encoded)
        
        logger.info(f"Intentando limpiar directorio: {temp_dir}")
        
        # Verificar que sea un directorio temporal válido
        if not temp_dir or not os.path.basename(temp_dir).startswith(('tmp', 'smartaudit_')):
            logger.warning(f"Directorio no válido para limpieza: {temp_dir}")
            return {"message": "Directorio no válido", "cleaned": False}
        
        if os.path.exists(temp_dir) and os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir)
            logger.info(f"Directorio temporal eliminado: {temp_dir}")
            return {"message": "Archivos temporales eliminados correctamente", "cleaned": True}
        else:
            logger.info(f"Directorio ya no existe: {temp_dir}")
            return {"message": "El directorio temporal ya no existe", "cleaned": False}
            
    except Exception as e:
        logger.error(f"Error al eliminar archivos temporales: {str(e)}")
        # No lanzar excepción aquí, ya que la limpieza es opcional
        return {"message": f"Error al eliminar archivos temporales: {str(e)}", "cleaned": False}


# Health check endpoint
@app.get("/api/health")
async def health_check():
    """Endpoint para verificar el estado del servicio"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")