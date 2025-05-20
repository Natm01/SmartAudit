# backend/app/services/file_processor.py
import os
import re
import csv
import pandas as pd
from typing import List, Dict, Any, Tuple
import codecs
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

def detect_encoding(file_path: str) -> str:
    """
    Detecta la codificación de un archivo.
    """
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with codecs.open(file_path, 'r', encoding=encoding) as f:
                f.read()
                return encoding
        except UnicodeDecodeError:
            continue
    
    # Default si no se puede detectar
    return 'utf-8'

def detect_separator(file_path: str, encoding: str) -> str:
    """
    Detecta el separador de un archivo CSV/TXT.
    """
    possible_delimiters = [';', ',', '\t', '|']
    
    try:
        with codecs.open(file_path, 'r', encoding=encoding) as f:
            sample = f.read(4096)
            
        counts = {sep: sample.count(sep) for sep in possible_delimiters}
        return max(counts.items(), key=lambda x: x[1])[0]
    except:
        # Si hay error, devolver separador por defecto
        return ';'

def parse_fixed_width_txt(file_path: str, encoding: str) -> List[Dict[str, Any]]:
    """
    Procesa un archivo TXT con formato de ancho fijo, específicamente para el formato de libro diario.
    """
    entries = []
    current_entry = None
    current_lines = []
    
    with codecs.open(file_path, 'r', encoding=encoding) as f:
        lines = f.readlines()
    
    i = 0
    while i < len(lines):
        line = lines[i].rstrip()
        
        # Ignorar líneas de separación o encabezados de tabla
        if line.startswith('---') or 'Nº act' in line or line.strip() == '':
            i += 1
            continue
        
        # Detectar línea de cabecera de asiento
        if re.match(r'^\d{8}', line):
            # Si ya tenemos una entrada, guardarla
            if current_entry:
                current_entry['lines'] = current_lines
                entries.append(current_entry)
                current_lines = []
            
            # Extraer datos de la cabecera
            parts = line.split()
            if len(parts) >= 5:
                entry_num = parts[0]
                doc_num = parts[1]
                fe_cont = parts[2]
                fe_doc = parts[3]
                ba_code = parts[4]
                
                # El texto de cabecera está después del código BA
                header_text = ' '.join(parts[5:]) if len(parts) > 5 else ""
                
                current_entry = {
                    'entry_number': entry_num,
                    'document_number': doc_num,
                    'accounting_date': fe_cont,
                    'doc_date': fe_doc,
                    'ba_code': ba_code,
                    'header_text': header_text
                }
            else:
                print(f"Error: Cabecera de asiento malformada: {line}")
                i += 1
                continue
        
        # Línea de detalle
        elif line.strip() and current_entry and line.startswith(' '):
            # Procesar primera columna (nombre de cuenta)
            account_name = line.strip()[:40].strip()
            
            # Extraer más información si la línea tiene suficiente longitud
            if len(line) > 40:
                # Procesamos los códigos y números de cuenta
                ap_code = line[40:48].strip()
                i_code = line[48:50].strip()
                mayus_code = line[50:65].strip() if len(line) > 65 else ""
                account_number = line[65:80].strip() if len(line) > 80 else ""
                
                # Improved amount parsing with better regex and error handling
                debit_ml = 0.0
                credit_ml = 0.0
                
                # Buscar importes en la línea con un regex mejorado
                # Manejar múltiples formatos: 1.234,56 o 1,234.56 o 1234.56 o 1234,56
                amount_pattern = r'(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})'
                amount_matches = list(re.finditer(amount_pattern, line))
                
                try:
                    # Si encontramos al menos un importe
                    if amount_matches:
                        # Para simplificar, asumimos que el primer importe es el DEBE y el segundo el HABER
                        # O si solo hay uno, su posición determina si es DEBE o HABER
                        if len(amount_matches) >= 2:
                            # Dos importes encontrados - primer columna es DEBE, segunda es HABER
                            debit_str = amount_matches[0].group(1)
                            credit_str = amount_matches[1].group(1)
                            
                            # Normalizar formato: reemplazar separadores y convertir a float
                            debit_ml = float(debit_str.replace('.', '').replace(',', '.'))
                            credit_ml = float(credit_str.replace('.', '').replace(',', '.'))
                        elif len(amount_matches) == 1:
                            # Un solo importe encontrado - determinar si es DEBE o HABER
                            amount_str = amount_matches[0].group(1)
                            amount_pos = amount_matches[0].start()
                            normalized_amount = float(amount_str.replace('.', '').replace(',', '.'))
                            
                            # Si la posición es anterior a la mitad de la línea, asumimos DEBE
                            # Si es posterior, asumimos HABER
                            middle_pos = len(line) // 2
                            if amount_pos < middle_pos:
                                debit_ml = normalized_amount
                            else:
                                credit_ml = normalized_amount
                except Exception as e:
                    print(f"Error parsing amounts in line: {line}, Error: {str(e)}")
                
                current_lines.append({
                    'account_name': account_name,
                    'ap_code': ap_code,
                    'i_code': i_code,
                    'mayus_code': mayus_code,
                    'account_number': account_number,
                    'debit': debit_ml,
                    'credit': credit_ml
                })
            else:
                # Si la línea es muy corta, podría ser la continuación de un nombre de cuenta
                # En ese caso, actualizamos el nombre de la última línea procesada
                if current_lines:
                    current_lines[-1]['account_name'] += ' ' + line.strip()
        
        i += 1
    
    # No olvidar la última entrada
    if current_entry:
        current_entry['lines'] = current_lines
        entries.append(current_entry)
    
    return entries

def parse_sumas_saldos_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Procesa un archivo Excel de sumas y saldos.
    """
    try:
        # Leer el archivo Excel
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        
        # Convertir a lista de listas
        data_rows = []
        for row in worksheet.iter_rows(values_only=True):
            data_rows.append(list(row))
        
        # Encontrar la fila de headers (buscar "Cta.mayor" o similar)
        header_row_idx = None
        for i, row in enumerate(data_rows):
            if row and any(cell and 'Cta.mayor' in str(cell) for cell in row):
                header_row_idx = i
                break
        
        if header_row_idx is None:
            raise ValueError("No se encontró la fila de headers en el archivo de sumas y saldos")
        
        # Mapeo de columnas basado en el archivo de ejemplo
        column_mapping = {
            'sociedad': 1,      # Soc.
            'cuenta': 3,        # Cta.mayor  
            'divisa': 11,       # Div.
            'descripcion': 6,   # Texto explicativo
            'arrastre': 12,     # Arrastre de saldos
            'saldoAnterior': 14, # Saldo per.anteriores  
            'debe': 16,         # Período de informe debe
            'haber': 17,        # Saldo Haber per.inf.
            'saldoAcumulado': 18 # Saldo acumulado
        }
        
        # Extraer datos desde después de los headers
        sumas_saldos_data = []
        for row_data in data_rows[header_row_idx + 2:]:  # +2 para saltar header y línea vacía
            if not row_data or len(row_data) <= column_mapping['cuenta']:
                continue
                
            # Verificar que hay una cuenta válida
            cuenta = row_data[column_mapping['cuenta']]
            if cuenta is None or cuenta == '':
                continue
                
            # Crear registro
            record = {
                'sociedad': str(row_data[column_mapping['sociedad']] or ''),
                'cuenta': str(cuenta),
                'descripcion': str(row_data[column_mapping['descripcion']] or ''),
                'moneda': str(row_data[10] or 'EUR'),  # Columna Mon.
                'divisa': str(row_data[column_mapping['divisa']] or ''),
                'arrastre': float(row_data[column_mapping['arrastre']] or 0),
                'saldoAnterior': float(row_data[column_mapping['saldoAnterior']] or 0),
                'debe': float(row_data[column_mapping['debe']] or 0),
                'haber': float(row_data[column_mapping['haber']] or 0),
                'saldoAcumulado': float(row_data[column_mapping['saldoAcumulado']] or 0)
            }
            
            sumas_saldos_data.append(record)
        
        return sumas_saldos_data
        
    except Exception as e:
        print(f"Error procesando archivo Excel de sumas y saldos: {str(e)}")
        raise

def parse_csv_sumas_saldos(file_path: str, encoding: str, separator: str) -> List[Dict[str, Any]]:
    """
    Procesa un archivo CSV de sumas y saldos.
    """
    try:
        df = pd.read_csv(file_path, encoding=encoding, sep=separator)
        
        # Buscar las columnas relevantes (nombres pueden variar)
        column_map = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            if 'cuenta' in col_lower or 'cta' in col_lower:
                column_map['cuenta'] = col
            elif 'sociedad' in col_lower or 'soc' in col_lower:
                column_map['sociedad'] = col
            elif 'descripcion' in col_lower or 'texto' in col_lower:
                column_map['descripcion'] = col
            elif 'arrastre' in col_lower:
                column_map['arrastre'] = col
            elif 'debe' in col_lower:
                column_map['debe'] = col
            elif 'haber' in col_lower:
                column_map['haber'] = col
            elif 'saldo' in col_lower and 'acumulado' in col_lower:
                column_map['saldoAcumulado'] = col
            elif 'moneda' in col_lower or 'mon' in col_lower:
                column_map['moneda'] = col
        
        # Procesar las filas
        records = []
        for _, row in df.iterrows():
            record = {
                'sociedad': str(row.get(column_map.get('sociedad', ''), '')),
                'cuenta': str(row.get(column_map.get('cuenta', ''), '')),
                'descripcion': str(row.get(column_map.get('descripcion', ''), '')),
                'moneda': str(row.get(column_map.get('moneda', ''), 'EUR')),
                'divisa': '',
                'arrastre': float(row.get(column_map.get('arrastre', ''), 0) or 0),
                'saldoAnterior': 0,
                'debe': float(row.get(column_map.get('debe', ''), 0) or 0),
                'haber': float(row.get(column_map.get('haber', ''), 0) or 0),
                'saldoAcumulado': float(row.get(column_map.get('saldoAcumulado', ''), 0) or 0)
            }
            
            # Solo agregar si tiene cuenta válida
            if record['cuenta']:
                records.append(record)
        
        return records
        
    except Exception as e:
        print(f"Error procesando archivo CSV de sumas y saldos: {str(e)}")
        raise

def process_libro_diario(temp_dir: str, files: List[str]) -> Dict[str, Any]:
    """
    Procesa archivos de libro diario.
    """
    all_entries = []
    
    for file_name in files:
        file_path = os.path.join(temp_dir, file_name)
        file_ext = os.path.splitext(file_name)[1].lower()
        
        encoding = detect_encoding(file_path)
        
        try:
            if file_ext in ['.csv', '.txt']:
                # Usar nuestra función mejorada para el análisis
                entries = parse_fixed_width_txt(file_path, encoding)
                all_entries.extend(entries)
            elif file_ext in ['.xlsx', '.xls']:
                # Procesar archivos Excel
                try:
                    df = pd.read_excel(file_path)
                    # Implementación para Excel pendiente
                    print(f"Procesando Excel: {file_name} - {df.shape}")
                except Exception as e:
                    print(f"Error al procesar Excel: {str(e)}")
        except Exception as e:
            print(f"Error procesando archivo {file_name}: {str(e)}")
    
    # Si no hay entradas, lanzar error
    if not all_entries:
        raise ValueError("No se encontraron entradas válidas en los archivos")
    
    # Extraer rangos de fecha
    try:
        accounting_dates = []
        doc_dates = []
        
        for entry in all_entries:
            try:
                # Extraer día, mes y año del formato DDMMYY
                fe_cont = entry["accounting_date"]
                fe_doc = entry["doc_date"]
                
                day_cont = int(fe_cont[:2])
                month_cont = int(fe_cont[2:4])
                year_short_cont = int(fe_cont[4:6])
                year_cont = 2000 + year_short_cont
                
                day_doc = int(fe_doc[:2])
                month_doc = int(fe_doc[2:4])
                year_short_doc = int(fe_doc[4:6])
                year_doc = 2000 + year_short_doc
                
                accounting_dates.append(datetime(year_cont, month_cont, day_cont))
                doc_dates.append(datetime(year_doc, month_doc, day_doc))
            except (ValueError, IndexError) as e:
                print(f"Error procesando fechas del asiento {entry.get('entry_number', 'desconocido')}: {str(e)}")
                continue
        
        accounting_date_range = ""
        registration_date_range = ""
        
        if accounting_dates:
            min_accounting = min(accounting_dates)
            max_accounting = max(accounting_dates)
            accounting_date_range = f"{min_accounting.strftime('%d/%m/%Y')} - {max_accounting.strftime('%d/%m/%Y')}"
        
        if doc_dates:
            min_doc = min(doc_dates)
            max_doc = max(doc_dates)
            registration_date_range = f"{min_doc.strftime('%d/%m/%Y')} - {max_doc.strftime('%d/%m/%Y')}"
    except Exception as e:
        print(f"Error calculando rangos de fecha: {str(e)}")
        accounting_date_range = "Desconocido"
        registration_date_range = "Desconocido"
    
    return {
        "entries": all_entries,
        "accounting_date_range": accounting_date_range,
        "registration_date_range": registration_date_range
    }

def process_sumas_saldos(temp_dir: str, files: List[str]) -> Dict[str, Any]:
    """
    Procesa archivos de sumas y saldos.
    """
    all_records = []
    
    for file_name in files:
        file_path = os.path.join(temp_dir, file_name)
        file_ext = os.path.splitext(file_name)[1].lower()
        
        try:
            if file_ext in ['.xlsx', '.xls']:
                records = parse_sumas_saldos_excel(file_path)
                all_records.extend(records)
            elif file_ext in ['.csv', '.txt']:
                # Para CSV/TXT de sumas y saldos
                encoding = detect_encoding(file_path)
                separator = detect_separator(file_path, encoding)
                records = parse_csv_sumas_saldos(file_path, encoding, separator)
                all_records.extend(records)
                
        except Exception as e:
            print(f"Error procesando archivo de sumas y saldos {file_name}: {str(e)}")
            raise
    
    if not all_records:
        raise ValueError("No se encontraron registros válidos en los archivos de sumas y saldos")
    
    # Calcular estadísticas
    total_debe = sum(record['debe'] for record in all_records)
    total_haber = sum(abs(record['haber']) for record in all_records)
    total_saldo = sum(record['saldoAcumulado'] for record in all_records)
    
    return {
        "records": all_records,
        "total_debe": total_debe,
        "total_haber": total_haber,
        "total_saldo": total_saldo,
        "cuentas_count": len(all_records)
    }